"""
Document Processor Service
Handles PDF, Word, Excel, Image, and URL processing
"""
import io
import json
import re
from typing import Optional

import httpx
from bs4 import BeautifulSoup
from PIL import Image

from src.utils.claude import chat, vision, get_sonnet_model

# File type detection
FILE_TYPE_MAP = {
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/msword": "doc",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xls",
    "text/csv": "csv",
    "image/png": "image",
    "image/jpeg": "image",
    "image/gif": "image",
    "image/webp": "image",
    "text/markdown": "markdown",
    "text/plain": "text",
}


class DocumentProcessor:
    """Process various document formats"""

    async def process_files(
        self, files: list[dict], user_id: str
    ) -> list[dict]:
        """
        Process multiple uploaded files

        Args:
            files: List of file dicts with 'name', 'content', 'mime_type'
            user_id: User ID

        Returns:
            List of processed results
        """
        results = []
        for file in files:
            try:
                result = await self.process_file(file)
                results.append(result)
            except Exception as e:
                results.append({
                    "file_name": file.get("name", "unknown"),
                    "file_type": "unknown",
                    "raw_content": "",
                    "error": str(e),
                })
        return results

    async def process_file(self, file: dict) -> dict:
        """
        Process a single file

        Args:
            file: File dict with 'name', 'content', 'mime_type'

        Returns:
            Processed result
        """
        name = file.get("name", "unknown")
        content = file.get("content", b"")
        mime_type = file.get("mime_type", "")

        # Detect file type
        file_type = self._detect_file_type(name, mime_type)

        # Extract raw content based on type
        raw_content = await self._extract_content(content, file_type, mime_type)

        # Use AI to understand and structure the content
        structured = await self._understand_content(raw_content, file_type, name)

        return {
            "file_name": name,
            "file_type": file_type,
            "raw_content": raw_content[:5000],  # Limit raw content
            "structured": structured,
        }

    async def process_url(self, url: str, user_id: str) -> dict:
        """
        Process a URL (fetch and extract content)

        Args:
            url: URL to process
            user_id: User ID

        Returns:
            Processed result
        """
        try:
            async with httpx.AsyncClient() as client:
                response = await client.get(url, follow_redirects=True, timeout=30)
                response.raise_for_status()

            content_type = response.headers.get("content-type", "")

            if "text/html" in content_type:
                # Parse HTML
                soup = BeautifulSoup(response.text, "html.parser")

                # Remove scripts and styles
                for tag in soup(["script", "style", "nav", "footer", "header"]):
                    tag.decompose()

                # Extract text
                raw_content = soup.get_text(separator="\n", strip=True)

                # Clean up
                raw_content = re.sub(r"\n{3,}", "\n\n", raw_content)
            else:
                raw_content = response.text

            # Use AI to understand
            structured = await self._understand_content(raw_content, "url", url)

            return {
                "file_name": url,
                "file_type": "url",
                "raw_content": raw_content[:5000],
                "structured": structured,
            }
        except Exception as e:
            return {
                "file_name": url,
                "file_type": "url",
                "raw_content": "",
                "error": str(e),
            }

    def _detect_file_type(self, name: str, mime_type: str) -> str:
        """Detect file type from name and MIME type"""
        # Check MIME type first
        if mime_type in FILE_TYPE_MAP:
            return FILE_TYPE_MAP[mime_type]

        # Check extension
        ext = name.lower().split(".")[-1] if "." in name else ""
        ext_map = {
            "pdf": "pdf",
            "docx": "docx",
            "doc": "doc",
            "xlsx": "xlsx",
            "xls": "xls",
            "csv": "csv",
            "png": "image",
            "jpg": "image",
            "jpeg": "image",
            "gif": "image",
            "webp": "image",
            "md": "markdown",
            "txt": "text",
        }
        return ext_map.get(ext, "unknown")

    async def _extract_content(
        self, content: bytes, file_type: str, mime_type: str
    ) -> str:
        """Extract text content from file"""
        if file_type == "pdf":
            return await self._extract_pdf(content)
        elif file_type in ("docx", "doc"):
            return await self._extract_docx(content)
        elif file_type in ("xlsx", "xls", "csv"):
            return await self._extract_excel(content, file_type)
        elif file_type == "image":
            return await self._extract_image(content, mime_type)
        elif file_type in ("markdown", "text"):
            return content.decode("utf-8", errors="ignore")
        else:
            return content.decode("utf-8", errors="ignore")

    async def _extract_pdf(self, content: bytes) -> str:
        """Extract text from PDF"""
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(content))
            text_parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            return "\n\n".join(text_parts)
        except Exception as e:
            return f"[PDF extraction error: {e}]"

    async def _extract_docx(self, content: bytes) -> str:
        """Extract text from DOCX"""
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            text_parts = []
            for para in doc.paragraphs:
                if para.text.strip():
                    text_parts.append(para.text)
            return "\n\n".join(text_parts)
        except Exception as e:
            return f"[DOCX extraction error: {e}]"

    async def _extract_excel(self, content: bytes, file_type: str) -> str:
        """Extract text from Excel/CSV"""
        try:
            if file_type == "csv":
                import csv
                text = content.decode("utf-8", errors="ignore")
                reader = csv.reader(io.StringIO(text))
                rows = list(reader)
                return "\n".join([", ".join(row) for row in rows[:100]])
            else:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content), read_only=True)
                text_parts = []
                for sheet in wb.worksheets[:3]:  # First 3 sheets
                    text_parts.append(f"## {sheet.title}")
                    for row in list(sheet.iter_rows(values_only=True))[:50]:
                        cells = [str(c) if c else "" for c in row]
                        text_parts.append(", ".join(cells))
                return "\n".join(text_parts)
        except Exception as e:
            return f"[Excel extraction error: {e}]"

    async def _extract_image(self, content: bytes, mime_type: str) -> str:
        """Extract content from image using Claude Vision"""
        try:
            prompt = """分析这张图片并提取以下信息:
1. 图片类型 (设计稿/截图/流程图/其他)
2. 主要内容描述
3. 如果是UI设计，列出所有UI元素
4. 如果有文字，提取所有文字内容
5. 如果是流程图，描述流程步骤

返回结构化的分析结果。"""

            result = await vision(content, mime_type, prompt)
            return result
        except Exception as e:
            return f"[Image extraction error: {e}]"

    async def _understand_content(
        self, content: str, file_type: str, file_name: str
    ) -> dict:
        """Use AI to understand and structure content"""
        if not content or len(content) < 10:
            return {
                "content_type": "unknown",
                "summary": "",
                "features": [],
                "ui_elements": [],
                "data_fields": [],
                "references": [],
            }

        prompt = f"""分析以下文档内容，提取产品需求信息。

文档类型: {file_type}
文件名: {file_name}

内容:
{content[:8000]}

请返回JSON格式的结构化分析:
{{
  "content_type": "requirement|design|data|reference|other",
  "summary": "内容摘要 (2-3句话)",
  "features": [
    {{"name": "功能名称", "description": "描述", "priority": "high|medium|low", "tags": ["标签"]}}
  ],
  "ui_elements": [
    {{"name": "元素名称", "type": "button|input|list|etc", "description": "描述"}}
  ],
  "data_fields": [
    {{"name": "字段名", "type": "string|number|etc", "description": "描述", "required": true}}
  ],
  "references": [
    {{"type": "url|document|image", "url": "链接", "description": "描述"}}
  ]
}}

只返回JSON。"""

        try:
            result = await chat(
                messages=[{"role": "user", "content": prompt}],
                model=get_sonnet_model(),
                max_tokens=2000,
            )

            # Parse JSON
            json_match = re.search(r"\{[\s\S]*\}", result)
            if json_match:
                return json.loads(json_match.group())
            return {
                "content_type": "unknown",
                "summary": result[:500],
                "features": [],
                "ui_elements": [],
                "data_fields": [],
                "references": [],
            }
        except Exception as e:
            return {
                "content_type": "error",
                "summary": str(e),
                "features": [],
                "ui_elements": [],
                "data_fields": [],
                "references": [],
            }
